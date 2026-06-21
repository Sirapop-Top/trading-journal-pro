import os
import json
import datetime
from typing import List, Optional, Dict
import pandas as pd
import yfinance as yf
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

app = FastAPI(title="Trading Journal API")

# Enable CORS for the frontend Vite development server
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify the allowed origin
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH = "db.json"
EXCEL_PATH = "Trading Journal.xlsx"

class Trade(BaseModel):
    id: str
    date: str
    portfolio: str
    assetName: str
    assetType: str
    currency: str
    action: str
    quantity: float
    priceUnit: float
    why: str
    remark: Optional[str] = ""
    feeRate: Optional[float] = 0.0

class PortfolioCreate(BaseModel):
    name: str

class PortfolioRename(BaseModel):
    oldName: str
    newName: str

class PositionTransfer(BaseModel):
    assetName: str
    sourcePortfolio: str
    targetPortfolio: str

class GoogleSheetSettings(BaseModel):
    google_sheet_id: str
    google_apps_script_url: Optional[str] = ""
    app_passcode: Optional[str] = ""

class TradeStrategyUpdate(BaseModel):
    why: str
    remark: Optional[str] = ""

class TradeTransfer(BaseModel):
    targetPortfolio: str



# In-memory cache for Yahoo Finance prices to prevent hitting limits and provide fast responses
price_cache = {
    "prices": {},
    "rates": {"THB": 1.0, "USD": 32.69, "EUR": 38.04}, # Default fallback rates
    "last_updated": None
}

# Mapping helper for tickers
TICKER_MAP = {
    "BJC": "BJC.BK",
    "KCE": "KCE.BK",
    "JMART": "JMART.BK",
    "ROJNA": "ROJNA.BK",
    "MSTR": "MSTR"
}

def get_ticker_symbol(asset_name: str, asset_type: str) -> str:
    # Check manual mapping first
    name_upper = asset_name.strip().upper()
    if name_upper in TICKER_MAP:
        return TICKER_MAP[name_upper]
    
    # If it is a Thai Stock and doesn't end in .BK, append it
    if asset_type.strip().lower() == "thai stock" and not name_upper.endswith(".BK"):
        return f"{name_upper}.BK"
        
    # If it is a Crypto and doesn't contain a dash (e.g. BTC-USD), append -USD
    if asset_type.strip().lower() == "crypto" and "-" not in name_upper:
        return f"{name_upper}-USD"
        
    return asset_name

def load_db():
    if os.path.exists(DB_PATH):
        try:
            with open(DB_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print("Error reading db.json, recreating...", e)
            
    # Fallback to Excel
    return load_from_excel_and_save()

def save_db(data):
    try:
        with open(DB_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print("Error saving db.json:", e)

def load_from_excel_and_save():
    if not os.path.exists(EXCEL_PATH):
        # Create default database if no Excel file exists
        default_data = {
            "trades": [],
            "portfolios": ["Main Investment", "Short-Term Trading", "Crypto"]
        }
        save_db(default_data)
        return default_data

    # Load existing settings to avoid overwriting them
    old_data = {}
    if os.path.exists(DB_PATH):
        try:
            with open(DB_PATH, "r", encoding="utf-8") as f:
                old_data = json.load(f)
        except Exception:
            pass

    try:
        print("Parsing Trading Journal.xlsx...")
        df = pd.read_excel(EXCEL_PATH, sheet_name="Journal")
        trades = []
        for index, row in df.iterrows():
            date_val = ""
            if pd.notna(row.get("Date")):
                # Format Date as YYYY-MM-DD
                if isinstance(row["Date"], (datetime.datetime, datetime.date)):
                    date_val = row["Date"].strftime("%Y-%m-%d")
                else:
                    date_val = str(row["Date"]).split(" ")[0]
            
            # Map columns safely
            asset_name = str(row.get("Asset Name", "")) if pd.notna(row.get("Asset Name")) else ""
            asset_type = str(row.get("Asset Type", "")) if pd.notna(row.get("Asset Type")) else ""
            currency = str(row.get("Currency", "THB")) if pd.notna(row.get("Currency")) else "THB"
            action = str(row.get("Action", "Buy")) if pd.notna(row.get("Action")) else "Buy"
            quantity = float(row.get("Quantity", 0.0)) if pd.notna(row.get("Quantity")) else 0.0
            price_unit = float(row.get("Price/Unit", 0.0)) if pd.notna(row.get("Price/Unit")) else 0.0
            why = str(row.get("Why (Decision Reason)", "")) if pd.notna(row.get("Why (Decision Reason)")) else ""
            remark = str(row.get("Remark", "")) if pd.notna(row.get("Remark")) and str(row.get("Remark")) != "nan" else ""
            
            # Default to Main Investment unless it looks like crypto or specified in Excel
            portfolio = str(row.get("Portfolio", "")) if pd.notna(row.get("Portfolio")) else ""
            if not portfolio:
                portfolio = "Main Investment"
                if asset_type.lower() == "crypto":
                    portfolio = "Crypto"

            # Parse fee rate from Excel
            fee_rate = 0.0
            if pd.notna(row.get("Fee Rate (%)")):
                try:
                    raw_fee = str(row.get("Fee Rate (%)")).replace("%", "").strip()
                    fee_rate = float(raw_fee)
                except Exception:
                    fee_rate = 0.0
            elif pd.notna(row.get("Fee Rate")):
                try:
                    raw_fee = str(row.get("Fee Rate")).replace("%", "").strip()
                    fee_rate = float(raw_fee)
                except Exception:
                    fee_rate = 0.0

            trade = {
                "id": str(index + 1),
                "date": date_val,
                "portfolio": portfolio,
                "assetName": asset_name,
                "assetType": asset_type,
                "currency": currency,
                "action": action,
                "quantity": quantity,
                "priceUnit": price_unit,
                "why": why,
                "remark": remark,
                "feeRate": fee_rate
            }
            trades.append(trade)

        # Retrieve unique portfolios from lists or default
        portfolios = ["Main Investment", "Short-Term Trading", "Crypto"]
        
        data = {
            "trades": trades,
            "portfolios": portfolios,
            "google_sheet_id": old_data.get("google_sheet_id", ""),
            "google_apps_script_url": old_data.get("google_apps_script_url", ""),
            "synced_google_form_timestamps": old_data.get("synced_google_form_timestamps", [])
        }
        save_db(data)
        print(f"Successfully loaded {len(trades)} trades from Excel sheet.")
        return data
    except Exception as e:
        print("Error parsing Excel:", e)
        # Fallback default keeping old settings if possible
        default_data = {
            "trades": [],
            "portfolios": ["Main Investment", "Short-Term Trading", "Crypto"],
            "google_sheet_id": old_data.get("google_sheet_id", ""),
            "google_apps_script_url": old_data.get("google_apps_script_url", ""),
            "synced_google_form_timestamps": old_data.get("synced_google_form_timestamps", [])
        }
        save_db(default_data)
        return default_data

def append_trade_to_excel(trade: Trade):
    """Safely appends a new trade row to the Excel sheet 'Journal' using openpyxl."""
    if not os.path.exists(EXCEL_PATH):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if 'Journal' not in wb.sheetnames:
            return
            
        sheet = wb['Journal']
        
        # We need to find the last row
        last_row = sheet.max_row
        new_row = last_row + 1
        
        # Excel columns: 
        # A: Date, B: Asset Name, C: Asset Type, D: Currency, E: Action, F: Quantity, G: Price/Unit, 
        # H: Amount, I: Current Price, J: Current Value, K: P&L, L: P&L %, M: Why (Decision Reason), N: Remark
        
        # We write values. H, J, K, L are formulas.
        # H: Quantity * Price/Unit -> =F{row}*G{row}
        # I: Current Price -> XLOOKUP from PriceDB or just temporary cell value
        # J: Quantity * Current Price -> =F{row}*I{row}
        # K: P&L -> =IF(E{row}="Buy",(I{row}-G{row})*F{row},(G{row}-I{row})*F{row})
        # L: P&L % -> =IF(H{row}=0,0,K{row}/H{row})
        
        date_obj = datetime.datetime.strptime(trade.date, "%Y-%m-%d") if trade.date else datetime.datetime.now()
        
        sheet.cell(row=new_row, column=1, value=date_obj)
        sheet.cell(row=new_row, column=2, value=trade.assetName)
        sheet.cell(row=new_row, column=3, value=trade.assetType)
        sheet.cell(row=new_row, column=4, value=trade.currency)
        sheet.cell(row=new_row, column=5, value=trade.action)
        sheet.cell(row=new_row, column=6, value=trade.quantity)
        sheet.cell(row=new_row, column=7, value=trade.priceUnit)
        
        # Formulas
        # Amount formula includes fee: Buy: qty * price * (1 + fee%/100); Sell: qty * price * (1 - fee%/100)
        sheet.cell(row=new_row, column=8, value=f'=IF(E{new_row}="Buy",F{new_row}*G{new_row}*(1+P{new_row}/100),F{new_row}*G{new_row}*(1-P{new_row}/100))')
        # Default Current Price to the buy price initially (will be updated when sheet is recalculated)
        sheet.cell(row=new_row, column=9, value=trade.priceUnit) 
        sheet.cell(row=new_row, column=10, value=f"=F{new_row}*I{new_row}")
        sheet.cell(row=new_row, column=11, value=f'=IF(E{new_row}="Buy",(I{new_row}-G{new_row})*F{new_row},(G{new_row}-I{new_row})*F{new_row})')
        sheet.cell(row=new_row, column=12, value=f"=IF(H{new_row}=0,0,K{new_row}/H{new_row})")
        
        sheet.cell(row=new_row, column=13, value=trade.why)
        sheet.cell(row=new_row, column=14, value=trade.remark or "")
        
        # Write portfolio
        if sheet.cell(row=1, column=15).value is None:
            sheet.cell(row=1, column=15, value="Portfolio")
        sheet.cell(row=new_row, column=15, value=trade.portfolio)

        # Write fee rate
        if sheet.cell(row=1, column=16).value is None:
            sheet.cell(row=1, column=16, value="Fee Rate (%)")
        sheet.cell(row=new_row, column=16, value=trade.feeRate)
        
        wb.save(EXCEL_PATH)
        print(f"Appended trade ID {trade.id} to Excel.")
    except Exception as e:
        print("Error appending trade to Excel:", e)

def fetch_live_market_data(assets: List[Dict[str, str]]):
    """Fetches stock prices and forex rates from Yahoo Finance."""
    tickers_to_fetch = set()
    
    # Currency pair tickers
    forex_tickers = ["USDTHB=X", "EURTHB=X"]
    for t in forex_tickers:
        tickers_to_fetch.add(t)
        
    # Map assets to yfinance tickers
    asset_to_ticker_map = {}
    for asset in assets:
        name = asset.get("name")
        asset_type = asset.get("type", "")
        if name:
            ticker = get_ticker_symbol(name, asset_type)
            asset_to_ticker_map[name] = ticker
            tickers_to_fetch.add(ticker)
            
    tickers_list = list(tickers_to_fetch)
    print(f"Fetching {len(tickers_list)} tickers from Yahoo Finance: {tickers_list}")
    
    fetched_prices = {}
    fetched_rates = {"THB": 1.0, "USD": 32.69, "EUR": 38.04} # defaults
    
    try:
        data = yf.download(tickers_list, period="1d", group_by="ticker", progress=False)
        
        for t in tickers_list:
            try:
                if t in data:
                    close_series = data[t]['Close'].dropna()
                    if not close_series.empty:
                        val = float(close_series.iloc[-1])
                        fetched_prices[t] = val
                    else:
                        fetched_prices[t] = None
                else:
                    fetched_prices[t] = None
            except Exception as inner_e:
                print(f"Error reading ticker {t}: {inner_e}")
                fetched_prices[t] = None
                
        # Set exchange rates
        if fetched_prices.get("USDTHB=X"):
            fetched_rates["USD"] = fetched_prices["USDTHB=X"]
        if fetched_prices.get("EURTHB=X"):
            fetched_rates["EUR"] = fetched_prices["EURTHB=X"]
            
        # Map tickers back to asset names
        mapped_prices = {}
        for asset_name, ticker in asset_to_ticker_map.items():
            if ticker in fetched_prices and fetched_prices[ticker] is not None:
                mapped_prices[asset_name] = fetched_prices[ticker]
            else:
                # Try fetching via Ticker history directly as a fallback
                try:
                    t_obj = yf.Ticker(ticker)
                    h = t_obj.history(period="1d")
                    if not h.empty:
                        val = float(h["Close"].iloc[-1])
                        mapped_prices[asset_name] = val
                        fetched_prices[ticker] = val
                    else:
                        mapped_prices[asset_name] = None
                except Exception:
                    mapped_prices[asset_name] = None
                    
        price_cache["prices"] = mapped_prices
        price_cache["rates"] = fetched_rates
        price_cache["last_updated"] = datetime.datetime.now().isoformat()
        
        print("Fetch completed. Rates:", fetched_rates)
        return True
    except Exception as e:
        print("Error fetching from Yahoo Finance:", e)
        return False

@app.get("/api/validate-ticker")
def validate_ticker(symbol: str, asset_type: str):
    symbol_stripped = symbol.strip()
    if not symbol_stripped:
        return {"valid": False, "message": "Ticker symbol cannot be empty."}
        
    ticker_symbol = get_ticker_symbol(symbol_stripped, asset_type)
    try:
        print(f"Validating ticker symbol: {ticker_symbol}")
        ticker = yf.Ticker(ticker_symbol)
        h = ticker.history(period="5d")
        if h.empty:
            return {
                "valid": False,
                "message": f"Ticker '{ticker_symbol}' not found or has no trading history on Yahoo Finance."
            }
        return {
            "valid": True,
            "ticker": ticker_symbol,
            "message": f"Ticker '{ticker_symbol}' is verified on Yahoo Finance."
        }
    except Exception as e:
        print(f"Validation connection error for {ticker_symbol}: {e}")
        return {
            "valid": True,
            "ticker": ticker_symbol,
            "message": "Yahoo Finance temporarily unreachable. Bypassed validation.",
            "warning": True
        }

@app.get("/api/data")
def get_dashboard_data():
    db_data = load_db()
    trades = db_data.get("trades", [])
    portfolios = db_data.get("portfolios", ["Main Investment", "Short-Term Trading", "Crypto"])
    
    # Identify unique assets from trades to fetch prices
    unique_assets = []
    seen_assets = set()
    for trade in trades:
        name = trade.get("assetName")
        if name and name not in seen_assets:
            seen_assets.add(name)
            unique_assets.append({
                "name": name,
                "type": trade.get("assetType", "")
            })
            
    # If price cache is empty, fetch prices synchronously on first load
    if not price_cache["prices"] or not price_cache["last_updated"]:
        fetch_live_market_data(unique_assets)
        
    return {
        "trades": trades,
        "portfolios": portfolios,
        "livePrices": price_cache["prices"],
        "liveRates": price_cache["rates"],
        "syncTime": price_cache["last_updated"],
        "fallbackUsed": not bool(price_cache["last_updated"])
    }

@app.post("/api/sync")
def sync_market_data():
    db_data = load_db()
    trades = db_data.get("trades", [])
    unique_assets = []
    seen_assets = set()
    for trade in trades:
        name = trade.get("assetName")
        if name and name not in seen_assets:
            seen_assets.add(name)
            unique_assets.append({
                "name": name,
                "type": trade.get("assetType", "")
            })
            
    success = fetch_live_market_data(unique_assets)
    if not success:
        # We return the cached values even if fetching failed
        return {
            "success": False,
            "message": "Yahoo Finance fetch failed, using cached/fallback data.",
            "livePrices": price_cache["prices"],
            "liveRates": price_cache["rates"],
            "syncTime": price_cache["last_updated"]
        }
        
    return {
        "success": True,
        "livePrices": price_cache["prices"],
        "liveRates": price_cache["rates"],
        "syncTime": price_cache["last_updated"]
    }

@app.post("/api/trades", response_model=Trade)
def add_trade(trade: Trade):
    db_data = load_db()
    trades = db_data.get("trades", [])
    
    # Check duplicate ID
    for t in trades:
        if t["id"] == trade.id:
            # Generate a new unique ID
            ids = [int(x["id"]) for x in trades if x["id"].isdigit()]
            trade.id = str(max(ids) + 1) if ids else "1"
            break
            
    new_trade_dict = trade.dict()
    trades.append(new_trade_dict)
    
    db_data["trades"] = trades
    save_db(db_data)
    
    # Also append to the Excel sheet
    append_trade_to_excel(trade)
    
    # Write to Google Sheets Apps Script if configured
    apps_script_url = db_data.get("google_apps_script_url")
    if apps_script_url:
        try:
            import requests
            print(f"[Google Apps Script Write] Logging trade to Google Sheet...")
            payload = {
                "date": trade.date,
                "portfolio": trade.portfolio,
                "assetName": trade.assetName,
                "assetType": trade.assetType,
                "currency": trade.currency,
                "action": trade.action,
                "quantity": trade.quantity,
                "priceUnit": trade.priceUnit,
                "why": trade.why,
                "remark": trade.remark,
                "feeRate": trade.feeRate
            }
            resp = requests.post(apps_script_url, json=payload, timeout=10)
            if resp.status_code == 200:
                print("[Google Apps Script Write] Trade logged successfully.")
                sync_google_sheet()
            else:
                print(f"[Google Apps Script Write] Failed with status code: {resp.status_code}")
        except Exception as e:
            print(f"[Google Apps Script Write] Error sending trade: {e}")
            
    # Invalidate cache if new asset is added
    if trade.assetName not in price_cache["prices"]:
        price_cache["prices"] = {} # clear cache so next load fetches everything including the new asset
        
    return trade

@app.delete("/api/trades/{trade_id}")
def delete_trade(trade_id: str):
    db_data = load_db()
    trades = db_data.get("trades", [])
    
    filtered_trades = [t for t in trades if t["id"] != trade_id]
    if len(filtered_trades) == len(trades):
        raise HTTPException(status_code=404, detail="Trade not found")
        
    db_data["trades"] = filtered_trades
    save_db(db_data)
    
    # Re-save to Excel to sync the deletion (or just update db.json)
    # Since Excel sheet is a historical log, syncing deletion might be optional, 
    # but let's update Excel too by rewriting the sheet or just keeping db.json in sync.
    # For now, updating db.json is enough, but let's try to overwrite Excel if it exists to keep them in sync.
    rewrite_excel_from_db(filtered_trades)
    
    return {"message": "Trade deleted successfully"}

@app.put("/api/trades/{trade_id}/strategy")
def update_trade_strategy(trade_id: str, update_data: TradeStrategyUpdate):
    db_data = load_db()
    trades = db_data.get("trades", [])
    
    found_trade = None
    for t in trades:
        if t["id"] == trade_id:
            t["why"] = update_data.why
            t["remark"] = update_data.remark if update_data.remark is not None else t.get("remark", "")
            found_trade = t
            break
            
    if not found_trade:
        raise HTTPException(status_code=404, detail="Trade not found")
        
    db_data["trades"] = trades
    save_db(db_data)
    
    # Rewrite Excel sheet to keep it in sync
    rewrite_excel_from_db(trades)
    
    # Write to Google Sheets Apps Script if configured (so cloud is updated too!)
    apps_script_url = db_data.get("google_apps_script_url")
    if apps_script_url:
        try:
            import requests
            print(f"[Google Apps Script Update] Updating trade strategy on Google Sheet...")
            payload = {
                "action": "updateTradeStrategy",
                "tradeId": trade_id,
                "why": update_data.why,
                "remark": update_data.remark
            }
            resp = requests.post(apps_script_url, json=payload, timeout=10)
            if resp.status_code == 200:
                print("[Google Apps Script Update] Trade strategy updated on Google Sheet successfully.")
                sync_google_sheet()
            else:
                print(f"[Google Apps Script Update] Failed with status code: {resp.status_code}")
        except Exception as e:
            print(f"[Google Apps Script Update] Error updating trade strategy: {e}")
            
    return found_trade

@app.post("/api/portfolios")
def add_portfolio(portfolio: PortfolioCreate):
    db_data = load_db()
    portfolios = db_data.get("portfolios", [])
    
    name = portfolio.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Portfolio name cannot be empty")
        
    if name in portfolios:
        raise HTTPException(status_code=400, detail="Portfolio already exists")
        
    portfolios.append(name)
    db_data["portfolios"] = portfolios
    save_db(db_data)
    return {"portfolios": portfolios}

@app.delete("/api/portfolios/{portfolio_name}")
def delete_portfolio(portfolio_name: str):
    db_data = load_db()
    portfolios = db_data.get("portfolios", [])
    trades = db_data.get("trades", [])
    
    if portfolio_name not in portfolios:
        raise HTTPException(status_code=404, detail="Portfolio not found")
        
    if portfolios and portfolio_name == portfolios[0]:
        raise HTTPException(status_code=400, detail="Cannot delete default portfolio")
        
    # Check if there are trades in this portfolio
    has_trades = any(t.get("portfolio") == portfolio_name for t in trades)
    if has_trades:
        raise HTTPException(status_code=400, detail="Cannot delete portfolio with existing trades. Please reassign or delete the trades first.")
        
    portfolios.remove(portfolio_name)
    db_data["portfolios"] = portfolios
    save_db(db_data)
    return {"portfolios": portfolios}

@app.put("/api/portfolios/rename")
def rename_portfolio(payload: PortfolioRename):
    db_data = load_db()
    portfolios = db_data.get("portfolios", [])
    trades = db_data.get("trades", [])
    
    old_name = payload.oldName.strip()
    new_name = payload.newName.strip()
    
    if not old_name or not new_name:
        raise HTTPException(status_code=400, detail="Portfolio names cannot be empty")
        
    if old_name not in portfolios:
        raise HTTPException(status_code=404, detail=f"Portfolio '{old_name}' not found")
        
    if new_name in portfolios:
        raise HTTPException(status_code=400, detail=f"Portfolio '{new_name}' already exists")
        
    # Rename in portfolios list
    idx = portfolios.index(old_name)
    portfolios[idx] = new_name
    
    # Migrate all connected trades in database
    updated_count = 0
    for trade in trades:
        if trade.get("portfolio") == old_name:
            trade["portfolio"] = new_name
            updated_count += 1
            
    db_data["portfolios"] = portfolios
    db_data["trades"] = trades
    save_db(db_data)
    
    # Sync with Excel sheet if any changes were made
    if updated_count > 0:
        rewrite_excel_from_db(trades)
        
    return {
        "portfolios": portfolios,
        "trades_updated": updated_count
    }

@app.put("/api/portfolios/transfer-position")
def transfer_position(payload: PositionTransfer):
    db_data = load_db()
    portfolios = db_data.get("portfolios", [])
    trades = db_data.get("trades", [])
    
    asset_name = payload.assetName.strip().upper()
    source = payload.sourcePortfolio.strip()
    target = payload.targetPortfolio.strip()
    
    if not asset_name or not source or not target:
        raise HTTPException(status_code=400, detail="Missing required transfer details")
        
    if source not in portfolios:
        raise HTTPException(status_code=404, detail=f"Source portfolio '{source}' not found")
        
    if target not in portfolios:
        raise HTTPException(status_code=404, detail=f"Target portfolio '{target}' not found")
        
    if source == target:
        raise HTTPException(status_code=400, detail="Source and target portfolios cannot be the same")
        
    # Migrate matching trades in database
    moved_count = 0
    for trade in trades:
        if trade.get("assetName", "").strip().upper() == asset_name and trade.get("portfolio") == source:
            trade["portfolio"] = target
            moved_count += 1
            
    if moved_count == 0:
        raise HTTPException(status_code=404, detail=f"No trades found for asset '{asset_name}' in portfolio '{source}'")
        
    db_data["trades"] = trades
    save_db(db_data)
    
    # Sync changes with Excel
    rewrite_excel_from_db(trades)
    
    return {
        "success": True,
        "trades_moved": moved_count,
        "message": f"Successfully transferred {asset_name} position ({moved_count} trades) from '{source}' to '{target}'."
    }

@app.put("/api/trades/{trade_id}/transfer")
def transfer_trade(trade_id: str, payload: TradeTransfer):
    db_data = load_db()
    portfolios = db_data.get("portfolios", [])
    trades = db_data.get("trades", [])
    
    target = payload.targetPortfolio.strip()
    if target not in portfolios:
        raise HTTPException(status_code=404, detail=f"Target portfolio '{target}' not found")
        
    found = False
    for trade in trades:
        if trade.get("id") == trade_id:
            trade["portfolio"] = target
            found = True
            break
            
    if not found:
        raise HTTPException(status_code=404, detail=f"Trade with ID '{trade_id}' not found")
        
    db_data["trades"] = trades
    save_db(db_data)
    rewrite_excel_from_db(trades)
    
    return {"success": True, "trade_id": trade_id, "new_portfolio": target}

# End of portfolio mappings API section

# Trigger initial load on startup
load_db()

def sync_google_sheet():
    db_data = load_db()
    sheet_id = db_data.get("google_sheet_id")
    if not sheet_id:
        print("[Google Sheet Sync] No google_sheet_id configured in db.json.")
        return False
        
    # Auto-extract Google Sheet ID if the user pasted the entire URL
    sheet_id = sheet_id.strip()
    if "docs.google.com/spreadsheets" in sheet_id:
        try:
            sheet_id = sheet_id.split("/d/")[1].split("/")[0]
            print(f"[Google Sheet Sync] Extracted clean Sheet ID: {sheet_id}")
        except Exception as parse_err:
            print(f"[Google Sheet Sync] Warning: Could not parse Sheet ID from URL: {parse_err}")
        
    try:
        import requests
        import io
        import pandas as pd
        
        print(f"[Google Sheet Sync] Syncing from Google Sheet: {sheet_id}...")
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
        response = requests.get(url, timeout=10)
        if response.status_code != 200:
            print(f"[Google Sheet Sync] Failed to download Google Sheet: HTTP {response.status_code}")
            return False
            
        df = pd.read_excel(io.BytesIO(response.content))
        
        synced_timestamps = db_data.get("synced_google_form_timestamps", [])
        synced_set = set(synced_timestamps)
        
        new_trades_count = 0
        trades_list = db_data.get("trades", [])
        
        # Build a set of existing local trade content signatures to prevent importing duplicates
        existing_local_sigs = set()
        for t in trades_list:
            d = t.get("date", "")
            a = t.get("assetName", "")
            act = t.get("action", "")
            q = t.get("quantity", 0)
            p = t.get("priceUnit", 0)
            sig = f"sig-{d}-{a}-{act.capitalize()}-{float(q)}-{float(p)}".strip()
            existing_local_sigs.add(sig)
        
        if df.empty:
            print("[Google Sheet Sync] Downloaded spreadsheet is empty. Ready for local-to-cloud initial sync.")
        else:
            cols = [str(c).strip().lower() for c in df.columns]
            
            def find_col_idx(keywords):
                for i, c in enumerate(cols):
                    if any(kw in c for kw in keywords):
                        return i
                return None
                
            timestamp_idx = find_col_idx(["timestamp"])
            date_idx = find_col_idx(["date"])
            asset_name_idx = find_col_idx(["asset name", "asset_name", "asset"])
            asset_type_idx = find_col_idx(["asset type", "asset_type", "type"])
            currency_idx = find_col_idx(["currency"])
            action_idx = find_col_idx(["action"])
            quantity_idx = find_col_idx(["quantity", "qty"])
            price_unit_idx = find_col_idx(["price/unit", "price_unit", "price unit", "price"])
            why_idx = find_col_idx(["why", "decision", "reason"])
            remark_idx = find_col_idx(["remark", "note"])
            fee_rate_idx = find_col_idx(["fee rate", "fee_rate", "fee %", "fee_pct", "fee"])
            
            for idx_row, row in df.iterrows():
                ts_val = row.iloc[timestamp_idx] if timestamp_idx is not None else None
                
                # If no timestamp column exists, construct a unique row signature to prevent duplicates
                if pd.isna(ts_val) or ts_val is None:
                    date_raw = row.iloc[date_idx] if date_idx is not None else ""
                    asset_raw = row.iloc[asset_name_idx] if asset_name_idx is not None else ""
                    action_raw = row.iloc[action_idx] if action_idx is not None else ""
                    qty_raw = row.iloc[quantity_idx] if quantity_idx is not None else 0
                    price_raw = row.iloc[price_unit_idx] if price_unit_idx is not None else 0
                    ts_str = f"sig-{date_raw}-{asset_raw}-{action_raw}-{qty_raw}-{price_raw}".strip()
                else:
                    ts_str = str(ts_val).strip()
                    
                # Parse date, asset, action, quantity, priceUnit to build the row content signature
                try:
                    row_date = ""
                    raw_date = row.iloc[date_idx] if date_idx is not None else datetime.datetime.now()
                    if pd.notna(raw_date):
                        if isinstance(raw_date, (datetime.datetime, datetime.date)):
                            row_date = raw_date.strftime("%Y-%m-%d")
                        else:
                            row_date = str(raw_date).split(" ")[0].strip()
                    else:
                        row_date = datetime.datetime.now().strftime("%Y-%m-%d")
                        
                    row_asset = str(row.iloc[asset_name_idx]).strip() if asset_name_idx is not None and pd.notna(row.iloc[asset_name_idx]) else ""
                    row_action = str(row.iloc[action_idx]).strip().capitalize() if action_idx is not None and pd.notna(row.iloc[action_idx]) else "Buy"
                    row_qty = float(row.iloc[quantity_idx]) if quantity_idx is not None and pd.notna(row.iloc[quantity_idx]) else 0.0
                    row_price = float(row.iloc[price_unit_idx]) if price_unit_idx is not None and pd.notna(row.iloc[price_unit_idx]) else 0.0
                    
                    row_sig = f"sig-{row_date}-{row_asset}-{row_action}-{row_qty}-{row_price}".strip()
                except Exception:
                    row_sig = ""
                    
                # Skip if already imported or matches an existing local trade signature
                if ts_str in synced_set or (row_sig and row_sig in existing_local_sigs):
                    if ts_str not in synced_set:
                        synced_timestamps.append(ts_str)
                        synced_set.add(ts_str)
                    continue
                    
                try:
                    date_val = ""
                    raw_date = row.iloc[date_idx] if date_idx is not None else datetime.datetime.now()
                    if pd.notna(raw_date):
                        if isinstance(raw_date, (datetime.datetime, datetime.date)):
                            date_val = raw_date.strftime("%Y-%m-%d")
                        else:
                            date_val = str(raw_date).split(" ")[0].strip()
                    else:
                        date_val = datetime.datetime.now().strftime("%Y-%m-%d")
                    
                    asset_name = str(row.iloc[asset_name_idx]).strip() if asset_name_idx is not None and pd.notna(row.iloc[asset_name_idx]) else ""
                    asset_type = str(row.iloc[asset_type_idx]).strip() if asset_type_idx is not None and pd.notna(row.iloc[asset_type_idx]) else "Thai Stock"
                    currency = str(row.iloc[currency_idx]).strip() if currency_idx is not None and pd.notna(row.iloc[currency_idx]) else "THB"
                    action = str(row.iloc[action_idx]).strip() if action_idx is not None and pd.notna(row.iloc[action_idx]) else "Buy"
                    
                    action = action.capitalize()
                    quantity = float(row.iloc[quantity_idx]) if quantity_idx is not None and pd.notna(row.iloc[quantity_idx]) else 0.0
                    price_unit = float(row.iloc[price_unit_idx]) if price_unit_idx is not None and pd.notna(row.iloc[price_unit_idx]) else 0.0
                    why = str(row.iloc[why_idx]).strip() if why_idx is not None and pd.notna(row.iloc[why_idx]) else ""
                    remark = str(row.iloc[remark_idx]).strip() if remark_idx is not None and pd.notna(row.iloc[remark_idx]) else ""
                    
                    fee_rate = 0.0
                    if fee_rate_idx is not None and pd.notna(row.iloc[fee_rate_idx]):
                        try:
                            raw_fee = str(row.iloc[fee_rate_idx]).replace("%", "").strip()
                            fee_rate = float(raw_fee)
                        except Exception:
                            fee_rate = 0.0

                    if not asset_name:
                        continue
                        
                    portfolios_list = db_data.get("portfolios", ["Main Trading", "BTC Stock", "Crypto"])
                    portfolio = portfolios_list[0] if portfolios_list else "Main Trading"
                    
                    if asset_type.lower() == "crypto" and "Crypto" in portfolios_list:
                        portfolio = "Crypto"
                    elif asset_type.lower() in ["global stock", "us stock"] and "BTC Stock" in portfolios_list:
                        portfolio = "BTC Stock"
                    
                    ids = [int(x["id"]) for x in trades_list if x["id"].isdigit()]
                    next_id = str(max(ids) + 1) if ids else "1"
                    
                    new_trade = {
                        "id": next_id,
                        "date": date_val,
                        "portfolio": portfolio,
                        "assetName": asset_name,
                        "assetType": asset_type,
                        "currency": currency,
                        "action": action,
                        "quantity": quantity,
                        "priceUnit": price_unit,
                        "why": why,
                        "remark": remark,
                        "feeRate": fee_rate
                    }
                    
                    trades_list.append(new_trade)
                    
                    trade_obj = Trade(
                        id=next_id,
                        date=date_val,
                        portfolio=portfolio,
                        assetName=asset_name,
                        assetType=asset_type,
                        currency=currency,
                        action=action,
                        quantity=quantity,
                        priceUnit=price_unit,
                        why=why,
                        remark=remark,
                        feeRate=fee_rate
                    )
                    append_trade_to_excel(trade_obj)
                    
                    synced_timestamps.append(ts_str)
                    synced_set.add(ts_str)
                    new_trades_count += 1
                    
                except Exception as row_err:
                    print(f"[Google Sheet Sync] Error parsing row {idx_row}: {row_err}")
                
        # Check if there are local trades that need to be uploaded to the cloud Google Sheet (auto-initialization)
        apps_script_url = db_data.get("google_apps_script_url")
        uploaded_count = 0
        if apps_script_url:
            local_trades = db_data.get("trades", [])
            
            for local_trade in local_trades:
                # Construct signature for local trade: sig-date-assetName-action-quantity-priceUnit
                date_val = local_trade.get("date", "")
                asset_val = local_trade.get("assetName", "")
                action_val = local_trade.get("action", "")
                qty_val = local_trade.get("quantity", 0)
                price_val = local_trade.get("priceUnit", 0)
                
                # Check signature
                sig = f"sig-{date_val}-{asset_val}-{action_val}-{qty_val}-{price_val}".strip()
                
                # If this local trade signature was not found in the Google Sheet, upload it!
                if sig not in synced_set:
                    print(f"[Google Sheet Sync] Uploading missing local trade to cloud sheet: {asset_val}...")
                    payload = {
                        "action": "addTrade",
                        "trade": {
                            "date": date_val,
                            "portfolio": local_trade.get("portfolio", "Main Trading"),
                            "assetName": asset_val,
                            "assetType": local_trade.get("assetType", "Thai Stock"),
                            "currency": local_trade.get("currency", "THB"),
                            "action": action_val,
                            "quantity": float(qty_val),
                            "priceUnit": float(price_val),
                            "why": local_trade.get("why", ""),
                            "remark": local_trade.get("remark", ""),
                            "timestamp": sig  # Pass the unique signature as timestamp so Google Sheets logs it
                        }
                    }
                    try:
                        resp = requests.post(apps_script_url, json=payload, timeout=10)
                        if resp.status_code == 200:
                            uploaded_count += 1
                            synced_set.add(sig)
                            synced_timestamps.append(sig)
                        else:
                            print(f"[Google Sheet Sync] Upload failed: HTTP {resp.status_code}")
                    except Exception as upload_err:
                        print(f"[Google Sheet Sync] Upload connection error: {upload_err}")
                        
            if uploaded_count > 0:
                print(f"[Google Sheet Sync] Successfully uploaded {uploaded_count} local trades to the cloud Google Sheet.")
        
        # Save imports and synced status
        if new_trades_count > 0 or uploaded_count > 0:
            db_data["trades"] = trades_list
            db_data["synced_google_form_timestamps"] = synced_timestamps
            save_db(db_data)
            print(f"[Google Sheet Sync] Sync databases updated successfully.")
            price_cache["prices"] = {}
        else:
            print("[Google Sheet Sync] No changes detected. Databases are in sync.")
            
        return True
    except Exception as e:
        print(f"[Google Sheet Sync] Error during synchronization: {e}")
        return False

@app.get("/api/google-sheet-settings")
def get_google_sheet_settings():
    db_data = load_db()
    return {
        "google_sheet_id": db_data.get("google_sheet_id", ""),
        "google_apps_script_url": db_data.get("google_apps_script_url", ""),
        "app_passcode": db_data.get("app_passcode", ""),
        "synced_count": len(db_data.get("trades", []))
    }

@app.post("/api/google-sheet-settings")
def save_google_sheet_settings(settings: GoogleSheetSettings):
    db_data = load_db()
    old_id = db_data.get("google_sheet_id", "")
    new_id = settings.google_sheet_id.strip()
    new_script_url = settings.google_apps_script_url.strip() if settings.google_apps_script_url else ""
    app_passcode = settings.app_passcode.strip() if settings.app_passcode else ""
    
    # Auto-extract Google Sheet ID if the user pasted the entire URL
    if "docs.google.com/spreadsheets" in new_id:
        try:
            new_id = new_id.split("/d/")[1].split("/")[0]
            print(f"[Google Sheet Settings] Extracted clean Sheet ID: {new_id}")
        except Exception as parse_err:
            print(f"[Google Sheet Settings] Warning: Could not parse Sheet ID from URL: {parse_err}")
            
    db_data["google_apps_script_url"] = new_script_url
    db_data["app_passcode"] = app_passcode
    
    if old_id != new_id:
        db_data["google_sheet_id"] = new_id
        db_data["synced_google_form_timestamps"] = []
    else:
        db_data["google_sheet_id"] = new_id
        
    save_db(db_data)
        
    sync_success = False
    if new_id:
        sync_success = sync_google_sheet()
        
    return {
        "success": True,
        "google_sheet_id": new_id,
        "google_apps_script_url": new_script_url,
        "app_passcode": app_passcode,
        "sync_success": sync_success
    }

@app.post("/api/google-sheet-sync")
def trigger_google_sheet_sync():
    db_data = load_db()
    sheet_id = db_data.get("google_sheet_id", "")
    if not sheet_id:
        raise HTTPException(status_code=400, detail="Google Sheet ID is not configured.")
        
    success = sync_google_sheet()
    if not success:
        raise HTTPException(status_code=500, detail="Failed to sync Google Sheet. Check log console.")
        
    db_data = load_db()
    return {
        "success": True,
        "trades": db_data.get("trades", [])
    }

def rewrite_excel_from_db(trades):
    if not os.path.exists(EXCEL_PATH):
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH)
        if 'Journal' not in wb.sheetnames:
            return
            
        sheet = wb['Journal']
        
        # Clear existing data rows (keep header row 1)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row)
            
        # Write trades one by one
        for i, t in enumerate(trades):
            row_num = i + 2
            date_str = t.get("date", "")
            try:
                date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d") if date_str else datetime.datetime.now()
            except Exception:
                date_obj = datetime.datetime.now()
                
            qty = float(t.get("quantity", 0))
            price = float(t.get("priceUnit", 0))
            
            sheet.cell(row=row_num, column=1, value=date_obj)
            sheet.cell(row=row_num, column=2, value=t.get("assetName", ""))
            sheet.cell(row=row_num, column=3, value=t.get("assetType", ""))
            sheet.cell(row=row_num, column=4, value=t.get("currency", "THB"))
            sheet.cell(row=row_num, column=5, value=t.get("action", "Buy"))
            sheet.cell(row=row_num, column=6, value=qty)
            sheet.cell(row=row_num, column=7, value=price)
            
            # Formulas
            # Amount formula includes fee: Buy: qty * price * (1 + fee%/100); Sell: qty * price * (1 - fee%/100)
            sheet.cell(row=row_num, column=8, value=f'=IF(E{row_num}="Buy",F{row_num}*G{row_num}*(1+P{row_num}/100),F{row_num}*G{row_num}*(1-P{row_num}/100))')
            sheet.cell(row=row_num, column=9, value=price) # default live price to cost
            sheet.cell(row=row_num, column=10, value=f"=F{row_num}*I{row_num}")
            sheet.cell(row=row_num, column=11, value=f'=IF(E{row_num}="Buy",(I{row_num}-G{row_num})*F{row_num},(G{row_num}-I{row_num})*F{row_num})')
            sheet.cell(row=row_num, column=12, value=f"=IF(H{row_num}=0,0,K{row_num}/H{row_num})")
            
            sheet.cell(row=row_num, column=13, value=t.get("why", ""))
            sheet.cell(row=row_num, column=14, value=t.get("remark", ""))
            
            # Write portfolio
            if sheet.cell(row=1, column=15).value is None:
                sheet.cell(row=1, column=15, value="Portfolio")
            sheet.cell(row=row_num, column=15, value=t.get("portfolio", "Main Investment"))

            # Write fee rate
            if sheet.cell(row=1, column=16).value is None:
                sheet.cell(row=1, column=16, value="Fee Rate (%)")
            sheet.cell(row=row_num, column=16, value=t.get("feeRate", 0.0))
            
        wb.save(EXCEL_PATH)
        print(f"[Cleanup] Re-wrote Excel sheet with {len(trades)} deduplicated trades.")
    except Exception as ex:
        print("Error rewriting Excel sheet:", ex)

def deduplicate_local_db():
    db_data = load_db()
    trades = db_data.get("trades", [])
    unique_trades = []
    seen_sigs = set()
    
    for t in trades:
        # Construct content signature
        date_val = t.get("date", "")
        asset_val = t.get("assetName", "")
        action_val = t.get("action", "")
        qty_val = t.get("quantity", 0)
        price_val = t.get("priceUnit", 0)
        sig = f"sig-{date_val}-{asset_val}-{action_val.capitalize()}-{float(qty_val)}-{float(price_val)}".strip()
        
        if sig not in seen_sigs:
            seen_sigs.add(sig)
            unique_trades.append(t)
            
    if len(unique_trades) < len(trades):
        print(f"[Cleanup] Deduplicated trades in db.json: from {len(trades)} to {len(unique_trades)}")
        db_data["trades"] = unique_trades
        save_db(db_data)
        rewrite_excel_from_db(unique_trades)

@app.on_event("startup")
def startup_event():
    print("[Startup] Checking for Google Sheet Mobile Sync configuration...")
    # Deduplicate local database and Excel on startup to clean up duplicates
    deduplicate_local_db()
    sync_google_sheet()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
